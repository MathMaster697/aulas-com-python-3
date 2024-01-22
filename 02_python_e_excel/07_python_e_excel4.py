# Importando a biblioteca openpyxl para manipulação de planilhas do Excel
import openpyxl
from openpyxl.styles import Alignment, Border, Side

# Abrindo uma planilha existente chamada 'alunos.xlsx'
planilha = openpyxl.load_workbook('alunos.xlsx')

# Selecionando uma folha específica da planilha, neste caso, a folha chamada 'Alunos'
folha = planilha['Alunos']

# Adicionando dez novos registros
for i in range(1, 11):
    # Criando dados fictícios para cada novo aluno
    nome = f'Novo Aluno {i}'
    idade = 20 + i
    curso = f'Curso {i}'

    # Encontrando a próxima linha vazia na coluna A
    linha_vazia = 2
    # Loop enquanto a célula na coluna A não estiver vazia
    while folha.cell(row=linha_vazia, column=1).value is not None:
        linha_vazia += 1

    # Escrevendo dados nas colunas A, B e C para o novo aluno na linha encontrada
    folha.cell(row=linha_vazia, column=1, value=nome)
    folha.cell(row=linha_vazia, column=2, value=idade)
    folha.cell(row=linha_vazia, column=3, value=curso)

    # Aplicando formatação
    for coluna in range(1, 4):
        # Centralizando o texto
        folha.cell(row=linha_vazia, column=coluna).alignment = Alignment(horizontal='center', vertical='center')
        # Adicionando bordas finas em todos os lados
        folha.cell(row=linha_vazia, column=coluna).border = Border(left=Side(style='thin'), 
                                                                     right=Side(style='thin'),
                                                                     top=Side(style='thin'),
                                                                     bottom=Side(style='thin'))

# Salvando as alterações na planilha com um novo nome 'alunos_atualizado3.xlsx'
planilha.save('alunos_atualizado3.xlsx')
