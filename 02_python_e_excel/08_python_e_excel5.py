# Importando a classe FPDF da biblioteca fpdf
from fpdf import FPDF # pip install fpdf

# Importando o módulo openpyxl e suas submódulos necessários
import openpyxl
from openpyxl.styles import Alignment, Border, Side

# Definindo uma classe chamada PDFCertificado que herda da classe FPDF
class PDFCertificado(FPDF):
    # Método para configurar o cabeçalho do PDF
    def header(self):
        # Configurando a fonte do texto no cabeçalho
        self.set_font('Arial', 'B', 12)
        # Adicionando uma célula centralizada com o título do certificado
        self.cell(0, 10, 'Certificado de Conclusão', 0, 1, 'C')

    # Método para configurar o rodapé do PDF
    def footer(self):
        # Movendo a posição Y para -15 unidades a partir do final da página
        self.set_y(-15)
        # Configurando a fonte do texto no rodapé
        self.set_font('Arial', 'I', 8)
        # Adicionando uma célula no rodapé com uma mensagem
        self.cell(0, 10, 'Esperamos que continue aprendendo e se desenvolvendo.', 0, 0, 'C')

# Definindo uma função chamada gerar_certificado
def gerar_certificado(nome, idade, curso):
    # Nome do arquivo PDF baseado no nome do aluno
    pdf_filename = f'Certificado_{nome}.pdf'

    # Criando instância da classe PDFCertificado
    pdf = PDFCertificado()
    # Adicionando uma nova página ao PDF
    pdf.add_page()

    # Adicionando conteúdo ao PDF
    # Configurando a fonte do texto para o título
    pdf.set_font('Arial', 'B', 16)
    # Adicionando uma célula centralizada com o título
    pdf.cell(0, 10, 'Certificado de Conclusão', 0, 1, 'C')

    # Configurando a fonte do texto para as informações do aluno
    pdf.set_font('Arial', '', 12)
    # Adicionando células com informações do aluno (Nome, Idade, Curso)
    pdf.cell(0, 10, f'Nome: {nome}', 0, 1, 'L')
    pdf.cell(0, 10, f'Idade: {idade}', 0, 1, 'L')
    pdf.cell(0, 10, f'Curso: {curso}', 0, 1, 'L')

    # Mensagem adicional no certificado
    mensagem = "Parabéns por concluir o curso com sucesso!\nEsperamos que continue aprendendo e se desenvolvendo."
    # Adicionando uma célula de múltiplas linhas com a mensagem adicional
    pdf.multi_cell(0, 10, mensagem)

    # Salvando o PDF com o nome específico
    pdf.output(pdf_filename)

# Abrindo uma planilha existente chamada 'alunos.xlsx'
planilha = openpyxl.load_workbook('alunos.xlsx')
# Selecionando a folha 'Alunos' na planilha
folha = planilha['Alunos']

# Iterando sobre os dados da planilha
for linha in folha.iter_rows(min_row=2, max_row=folha.max_row, values_only=True):
    # Desempacotando os valores da linha (nome, idade, curso)
    nome, idade, curso = linha
    # Gerando certificado para cada aluno
    gerar_certificado(nome, idade, curso)

# Exibindo mensagem indicando que os certificados foram gerados com sucesso
print("Certificados gerados com sucesso!")
